"""
批量邮件发送工具 - 企业微信邮箱版
通过企业微信邮箱 SMTP 批量发送个性化邮件
支持通过 IMAP 拉取收件箱并按日期导出为 Excel
"""

import os
import json
import smtplib
import imaplib
import email as email_mod
import time
import uuid
import mimetypes
import traceback
import re
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.header import Header, decode_header
from email.utils import parsedate_to_datetime, parseaddr
from email import encoders
from pathlib import Path

import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from dotenv import load_dotenv

# 加载 .env 文件中的环境变量
load_dotenv(Path(__file__).parent / ".env")

app = Flask(__name__, static_folder="static")
CORS(app)

UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

ATTACH_DIR = Path(__file__).parent / "attachments"
ATTACH_DIR.mkdir(exist_ok=True)

EXPORT_DIR = Path(__file__).parent / "exports"
EXPORT_DIR.mkdir(exist_ok=True)

# 全局发送状态追踪
send_tasks = {}

# 全局收件箱导出任务追踪
inbox_tasks = {}


# ── 静态页面 ──────────────────────────────────────────────────────────────────

@app.route("/api/defaults")
def get_defaults():
    """返回环境变量中的默认 SMTP 配置，前端自动填充"""
    return jsonify({
        "smtp_host": os.environ.get("SMTP_HOST", ""),
        "smtp_port": os.environ.get("SMTP_PORT", "465"),
        "email": os.environ.get("SENDER_EMAIL", ""),
        "password_set": bool(os.environ.get("SENDER_PASSWORD")),
        "sender_name": os.environ.get("SENDER_NAME", "CSDN"),
        "imap_host": os.environ.get("IMAP_HOST", "imap.exmail.qq.com"),
        "imap_port": os.environ.get("IMAP_PORT", "993"),
    })


@app.route("/")
def index():
    return send_from_directory("static", "index.html")


# ── 上传并解析 Excel ──────────────────────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
def upload_excel():
    """上传 Excel 文件，返回列名和前几行预览数据"""
    if "file" not in request.files:
        return jsonify({"error": "未上传文件"}), 400

    f = request.files["file"]
    filename = f.filename
    if not filename:
        return jsonify({"error": "文件名为空"}), 400

    ext = Path(filename).suffix.lower()
    if ext not in (".xls", ".xlsx", ".csv"):
        return jsonify({"error": "仅支持 .xls / .xlsx / .csv 格式"}), 400

    save_path = UPLOAD_DIR / filename
    f.save(str(save_path))

    try:
        headers, rows = parse_excel(save_path)
    except Exception as e:
        return jsonify({"error": f"解析失败: {e}"}), 400

    return jsonify({
        "filename": filename,
        "headers": headers,
        "preview": rows[:10],
        "totalRows": len(rows),
    })


@app.route("/api/parse-all", methods=["POST"])
def parse_all_rows():
    """返回 Excel 全部数据行"""
    data = request.json
    filename = data.get("filename")
    if not filename:
        return jsonify({"error": "缺少 filename"}), 400

    save_path = UPLOAD_DIR / filename
    if not save_path.exists():
        return jsonify({"error": "文件不存在，请重新上传"}), 404

    try:
        headers, rows = parse_excel(save_path)
    except Exception as e:
        return jsonify({"error": f"解析失败: {e}"}), 400

    return jsonify({"headers": headers, "rows": rows})


# ── 附件管理 ──────────────────────────────────────────────────────────────────

@app.route("/api/upload-attachments", methods=["POST"])
def upload_attachments():
    """批量上传附件文件"""
    if "files" not in request.files:
        return jsonify({"error": "未上传文件"}), 400

    files = request.files.getlist("files")
    saved = []
    for f in files:
        if f.filename:
            save_path = ATTACH_DIR / f.filename
            f.save(str(save_path))
            saved.append(f.filename)

    return jsonify({
        "uploaded": len(saved),
        "files": saved,
        "all_files": _list_attachments(),
    })


@app.route("/api/attachments")
def list_attachments():
    """列出已上传的所有附件"""
    return jsonify({"files": _list_attachments()})


@app.route("/api/attachments/<filename>", methods=["DELETE"])
def delete_attachment(filename):
    """删除单个附件"""
    fpath = ATTACH_DIR / filename
    if fpath.exists():
        fpath.unlink()
    return jsonify({"success": True, "files": _list_attachments()})


@app.route("/api/clear-attachments", methods=["POST"])
def clear_attachments():
    """清空所有附件"""
    for f in ATTACH_DIR.iterdir():
        if f.is_file() and not f.name.startswith("."):
            f.unlink()
    return jsonify({"success": True, "files": []})


def _list_attachments():
    """返回附件目录中的文件列表"""
    files = []
    for f in sorted(ATTACH_DIR.iterdir()):
        if f.is_file() and not f.name.startswith("."):
            size = f.stat().st_size
            if size < 1024:
                size_str = f"{size} B"
            elif size < 1024 * 1024:
                size_str = f"{size / 1024:.1f} KB"
            else:
                size_str = f"{size / (1024 * 1024):.1f} MB"
            files.append({"name": f.name, "size": size_str, "size_bytes": size})
    return files


# ── 发送邮件 ──────────────────────────────────────────────────────────────────

@app.route("/api/send", methods=["POST"])
def send_emails():
    """
    批量发送邮件
    请求体:
    {
        "smtp_host": "smtp.exmail.qq.com",
        "smtp_port": 465,
        "email": "xxx@csdn.net",
        "password": "xxx",
        "subject": "邮件主题，支持 {{姓名}} 变量",
        "body_html": "<p>正文 HTML，支持 {{姓名}} 变量</p>",
        "recipients": [
            {"邮箱": "a@b.com", "姓名": "张三", ...},
            ...
        ],
        "email_col": "邮箱",
        "delay": 2
    }
    """
    data = request.json
    smtp_host = data.get("smtp_host") or os.environ.get("SMTP_HOST", "smtp.exmail.qq.com")
    smtp_port = int(data.get("smtp_port") or os.environ.get("SMTP_PORT", 465))
    sender_email = data.get("email") or os.environ.get("SENDER_EMAIL", "")
    sender_password = data.get("password") or os.environ.get("SENDER_PASSWORD", "")
    subject_tpl = data.get("subject", "")
    body_tpl = data.get("body_html", "")
    recipients = data.get("recipients", [])
    email_col = data.get("email_col", "邮箱")
    delay = float(data.get("delay", 2))
    sender_name = data.get("sender_name") or os.environ.get("SENDER_NAME", "CSDN")
    attach_col = data.get("attach_col", "")  # 附件文件名列（可选）

    if not sender_email or not sender_password:
        return jsonify({"error": "请填写发件人邮箱和密码"}), 400
    if not recipients:
        return jsonify({"error": "收件人列表为空"}), 400
    if not subject_tpl:
        return jsonify({"error": "请填写邮件主题"}), 400

    task_id = str(uuid.uuid4())[:8]
    send_tasks[task_id] = {
        "total": len(recipients),
        "sent": 0,
        "failed": 0,
        "results": [],
        "status": "running",
    }

    # 在后台线程中发送
    import threading
    t = threading.Thread(
        target=_send_worker,
        args=(task_id, smtp_host, smtp_port, sender_email, sender_password,
              sender_name, subject_tpl, body_tpl, recipients, email_col,
              attach_col, delay),
        daemon=True,
    )
    t.start()

    return jsonify({"task_id": task_id, "total": len(recipients)})


def _send_worker(task_id, smtp_host, smtp_port, sender_email, sender_password,
                 sender_name, subject_tpl, body_tpl, recipients, email_col,
                 attach_col, delay):
    """后台发送线程"""
    task = send_tasks[task_id]
    server = None

    try:
        # 连接 SMTP
        if smtp_port == 465:
            server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
            server.starttls()

        server.login(sender_email, sender_password)

        for i, recipient in enumerate(recipients):
            to_email = recipient.get(email_col, "").strip()
            if not to_email or "@" not in to_email:
                task["results"].append({
                    "index": i,
                    "email": to_email,
                    "status": "skipped",
                    "message": "邮箱地址无效",
                })
                task["failed"] += 1
                continue

            # 替换模板变量
            subject = _render_template(subject_tpl, recipient)
            body = _render_template(body_tpl, recipient)

            # 查找附件文件
            attach_files = []
            if attach_col:
                raw_names = recipient.get(attach_col, "").strip()
                if raw_names:
                    # 支持多个附件，用分号或逗号分隔
                    for name in raw_names.replace("；", ";").replace("，", ",").replace(";", ",").split(","):
                        name = name.strip()
                        if not name:
                            continue
                        fpath = ATTACH_DIR / name
                        if fpath.exists() and fpath.is_file():
                            attach_files.append(fpath)
                        else:
                            task["results"].append({
                                "index": i,
                                "email": to_email,
                                "status": "failed",
                                "message": f"附件未找到: {name}",
                            })
                            task["failed"] += 1
                            break
                    else:
                        # 所有附件都找到了，继续发送
                        pass

                    # 如果上面 break 了（有附件没找到），跳过这个收件人
                    if any(r["index"] == i and r["status"] == "failed" for r in task["results"]):
                        continue

            try:
                msg = MIMEMultipart("mixed")
                msg["From"] = f"{sender_name} <{sender_email}>"
                msg["To"] = to_email
                msg["Subject"] = Header(subject, "utf-8")

                # HTML 正文
                html_part = MIMEText(body, "html", "utf-8")
                msg.attach(html_part)

                # 添加附件
                for fpath in attach_files:
                    ctype, encoding = mimetypes.guess_type(str(fpath))
                    if ctype is None:
                        ctype = "application/octet-stream"
                    maintype, subtype = ctype.split("/", 1)

                    with open(fpath, "rb") as af:
                        part = MIMEBase(maintype, subtype)
                        part.set_payload(af.read())
                    encoders.encode_base64(part)
                    # 用 RFC 2231 编码中文文件名
                    part.add_header(
                        "Content-Disposition", "attachment",
                        filename=("utf-8", "", fpath.name),
                    )
                    msg.attach(part)

                server.sendmail(sender_email, [to_email], msg.as_string())

                attach_info = f"（附件: {', '.join(f.name for f in attach_files)}）" if attach_files else ""
                task["results"].append({
                    "index": i,
                    "email": to_email,
                    "status": "success",
                    "message": f"发送成功{attach_info}",
                })
                task["sent"] += 1

            except Exception as e:
                task["results"].append({
                    "index": i,
                    "email": to_email,
                    "status": "failed",
                    "message": str(e),
                })
                task["failed"] += 1

            # 发送间隔，避免被限流
            if i < len(recipients) - 1:
                time.sleep(delay)

    except Exception as e:
        task["results"].append({
            "index": -1,
            "email": "",
            "status": "error",
            "message": f"SMTP 连接/登录失败: {e}\n{traceback.format_exc()}",
        })
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass
        task["status"] = "done"


@app.route("/api/send-status/<task_id>")
def send_status(task_id):
    """查询发送进度"""
    task = send_tasks.get(task_id)
    if not task:
        return jsonify({"error": "任务不存在"}), 404
    return jsonify(task)


@app.route("/api/test-smtp", methods=["POST"])
def test_smtp():
    """测试 SMTP 连接"""
    data = request.json
    smtp_host = data.get("smtp_host") or os.environ.get("SMTP_HOST", "smtp.exmail.qq.com")
    smtp_port = int(data.get("smtp_port") or os.environ.get("SMTP_PORT", 465))
    email = data.get("email") or os.environ.get("SENDER_EMAIL", "")
    password = data.get("password") or os.environ.get("SENDER_PASSWORD", "")

    if not email or not password:
        return jsonify({"error": "请填写邮箱和密码"}), 400

    try:
        if smtp_port == 465:
            server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=15)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=15)
            server.starttls()

        server.login(email, password)
        server.quit()
        return jsonify({"success": True, "message": "SMTP 连接成功！"})
    except Exception as e:
        return jsonify({"success": False, "message": f"连接失败: {e}"})


# ── 收件箱导出 (IMAP) ────────────────────────────────────────────────────────

@app.route("/api/test-imap", methods=["POST"])
def test_imap():
    """测试 IMAP 连接"""
    data = request.json or {}
    imap_host = data.get("imap_host") or os.environ.get("IMAP_HOST", "imap.exmail.qq.com")
    imap_port = int(data.get("imap_port") or os.environ.get("IMAP_PORT", 993))
    email_addr = data.get("email") or os.environ.get("SENDER_EMAIL", "")
    password = data.get("password") or os.environ.get("SENDER_PASSWORD", "")

    if not email_addr or not password:
        return jsonify({"success": False, "message": "请填写邮箱和密码"}), 400

    try:
        imap = imaplib.IMAP4_SSL(imap_host, imap_port, timeout=15)
        imap.login(email_addr, password)
        imap.select("INBOX", readonly=True)
        imap.logout()
        return jsonify({"success": True, "message": "IMAP 连接成功"})
    except Exception as e:
        return jsonify({"success": False, "message": f"连接失败: {e}"}), 200


@app.route("/api/export-inbox", methods=["POST"])
def export_inbox():
    """按日期范围拉取收件箱邮件并导出 Excel

    请求体:
    {
        "imap_host": "imap.exmail.qq.com",
        "imap_port": 993,
        "email": "xxx@csdn.net",
        "password": "xxx",
        "start_date": "2026-05-01",    # 可选，默认近 7 天
        "end_date": "2026-05-09",      # 可选，含当日
        "folder": "INBOX",             # 可选
        "include_body": true            # 是否包含正文内容
    }
    """
    data = request.json or {}
    imap_host = data.get("imap_host") or os.environ.get("IMAP_HOST", "imap.exmail.qq.com")
    imap_port = int(data.get("imap_port") or os.environ.get("IMAP_PORT", 993))
    email_addr = data.get("email") or os.environ.get("SENDER_EMAIL", "")
    password = data.get("password") or os.environ.get("SENDER_PASSWORD", "")
    start_date = data.get("start_date", "").strip()
    end_date = data.get("end_date", "").strip()
    folder = data.get("folder", "INBOX") or "INBOX"
    include_body = data.get("include_body", True)

    if not email_addr or not password:
        return jsonify({"error": "请填写邮箱和密码"}), 400

    # 日期默认值：近 7 天
    try:
        if end_date:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        else:
            end_dt = datetime.now()
        if start_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        else:
            start_dt = end_dt - timedelta(days=7)
    except ValueError as e:
        return jsonify({"error": f"日期格式错误，应为 YYYY-MM-DD: {e}"}), 400

    if start_dt > end_dt:
        return jsonify({"error": "开始日期不能晚于结束日期"}), 400

    task_id = str(uuid.uuid4())[:8]
    inbox_tasks[task_id] = {
        "status": "running",
        "total": 0,
        "fetched": 0,
        "message": "正在连接 IMAP...",
        "file": None,
        "error": None,
    }

    import threading
    t = threading.Thread(
        target=_inbox_export_worker,
        args=(task_id, imap_host, imap_port, email_addr, password,
              start_dt, end_dt, folder, include_body),
        daemon=True,
    )
    t.start()

    return jsonify({
        "task_id": task_id,
        "start_date": start_dt.strftime("%Y-%m-%d"),
        "end_date": end_dt.strftime("%Y-%m-%d"),
    })


@app.route("/api/export-status/<task_id>")
def export_status(task_id):
    """查询导出进度"""
    task = inbox_tasks.get(task_id)
    if not task:
        return jsonify({"error": "任务不存在"}), 404
    return jsonify(task)


@app.route("/api/download-export/<task_id>")
def download_export(task_id):
    """下载导出的 Excel 文件"""
    task = inbox_tasks.get(task_id)
    if not task or not task.get("file"):
        return jsonify({"error": "文件不存在"}), 404
    fpath = EXPORT_DIR / task["file"]
    if not fpath.exists():
        return jsonify({"error": "文件已被清理"}), 404
    return send_file(str(fpath), as_attachment=True, download_name=task["file"])


def _inbox_export_worker(task_id, imap_host, imap_port, email_addr, password,
                         start_dt, end_dt, folder, include_body):
    """后台 IMAP 拉取线程"""
    task = inbox_tasks[task_id]
    imap = None
    try:
        task["message"] = f"连接 {imap_host}:{imap_port} ..."
        imap = imaplib.IMAP4_SSL(imap_host, imap_port, timeout=30)
        imap.login(email_addr, password)

        # 选择文件夹（有的服务商文件夹名为中文，用引号包裹）
        try:
            status, _ = imap.select(folder, readonly=True)
        except Exception:
            status = "NO"
        if status != "OK":
            status, _ = imap.select(f'"{folder}"', readonly=True)
        if status != "OK":
            raise RuntimeError(f"无法打开文件夹: {folder}")

        # 使用 IMAP SINCE/BEFORE 过滤；BEFORE 是不含当日，所以 +1
        since_str = start_dt.strftime("%d-%b-%Y")
        before_str = (end_dt + timedelta(days=1)).strftime("%d-%b-%Y")
        task["message"] = f"搜索 {since_str} 到 {end_dt.strftime('%d-%b-%Y')} 的邮件..."
        status, search_data = imap.search(None, "SINCE", since_str, "BEFORE", before_str)
        if status != "OK":
            raise RuntimeError(f"搜索失败: {status}")

        ids = search_data[0].split() if search_data and search_data[0] else []
        task["total"] = len(ids)
        task["message"] = f"共找到 {len(ids)} 封邮件，开始拉取..."

        emails = []
        for idx, num in enumerate(ids, 1):
            try:
                status, msg_data = imap.fetch(num, "(RFC822)")
                if status != "OK" or not msg_data or not msg_data[0]:
                    continue
                raw = msg_data[0][1]
                msg = email_mod.message_from_bytes(raw)

                date_obj = _parse_email_date(msg.get("Date", ""))
                # 再次按日期范围过滤（IMAP 服务器时区偏差兜底）
                if date_obj:
                    if date_obj.date() < start_dt.date() or date_obj.date() > end_dt.date():
                        task["fetched"] = idx
                        continue

                from_raw = msg.get("From", "")
                from_name, from_email = parseaddr(from_raw)
                from_name = _decode_mime_str(from_name)
                subject = _decode_mime_str(msg.get("Subject", ""))
                body = _extract_body(msg) if include_body else ""

                emails.append({
                    "date": date_obj.strftime("%Y-%m-%d %H:%M:%S") if date_obj else "",
                    "date_only": date_obj.strftime("%Y-%m-%d") if date_obj else "",
                    "from_email": from_email,
                    "from_name": from_name,
                    "subject": subject,
                    "body": body,
                })
            except Exception as e:
                task["message"] = f"第 {idx} 封邮件解析失败: {e}"
            task["fetched"] = idx

        # 按日期升序排列
        emails.sort(key=lambda x: x["date"])

        task["message"] = f"正在生成 Excel（{len(emails)} 条记录）..."
        filename = f"inbox_{start_dt.strftime('%Y%m%d')}-{end_dt.strftime('%Y%m%d')}_{task_id}.xlsx"
        out_path = EXPORT_DIR / filename
        _write_inbox_xlsx(out_path, emails)

        task["file"] = filename
        task["count"] = len(emails)
        task["status"] = "done"
        task["message"] = f"导出完成，共 {len(emails)} 条"

    except Exception as e:
        task["status"] = "error"
        task["error"] = str(e)
        task["message"] = f"导出失败: {e}"
    finally:
        if imap is not None:
            try:
                imap.logout()
            except Exception:
                pass


def _decode_mime_str(s):
    """解码邮件头中的 MIME 编码字符串（如 =?UTF-8?B?...?=）"""
    if not s:
        return ""
    try:
        parts = decode_header(s)
        out = []
        for text, charset in parts:
            if isinstance(text, bytes):
                try:
                    out.append(text.decode(charset or "utf-8", errors="replace"))
                except LookupError:
                    out.append(text.decode("utf-8", errors="replace"))
            else:
                out.append(text)
        return "".join(out).strip()
    except Exception:
        return str(s)


def _parse_email_date(date_str):
    """解析邮件 Date 头为 datetime 对象"""
    if not date_str:
        return None
    try:
        dt = parsedate_to_datetime(date_str)
        # 转为本地时间（去除时区信息以便统一比较）
        if dt.tzinfo is not None:
            dt = dt.astimezone().replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _extract_body(msg):
    """从 email.message.Message 提取正文文本（优先纯文本，其次 HTML 去标签）"""
    text_body = ""
    html_body = ""

    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            disp = str(part.get("Content-Disposition", ""))
            if "attachment" in disp.lower():
                continue
            if ctype == "text/plain" and not text_body:
                text_body = _decode_payload(part)
            elif ctype == "text/html" and not html_body:
                html_body = _decode_payload(part)
    else:
        ctype = msg.get_content_type()
        if ctype == "text/plain":
            text_body = _decode_payload(msg)
        elif ctype == "text/html":
            html_body = _decode_payload(msg)

    if text_body.strip():
        return text_body.strip()
    if html_body.strip():
        return _html_to_text(html_body).strip()
    return ""


def _decode_payload(part):
    """解码邮件正文 payload"""
    try:
        payload = part.get_payload(decode=True)
        if payload is None:
            return ""
        charset = part.get_content_charset() or "utf-8"
        try:
            return payload.decode(charset, errors="replace")
        except LookupError:
            return payload.decode("utf-8", errors="replace")
    except Exception:
        return ""


def _html_to_text(html):
    """简单的 HTML 转纯文本（去标签、还原常用实体）"""
    text = re.sub(r"<br\s*/?>", "\n", html, flags=re.IGNORECASE)
    text = re.sub(r"</p>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = (text
            .replace("&nbsp;", " ")
            .replace("&amp;", "&")
            .replace("&lt;", "<")
            .replace("&gt;", ">")
            .replace("&quot;", '"')
            .replace("&#39;", "'"))
    text = re.sub(r"\n\s*\n", "\n\n", text)
    return text


def _write_inbox_xlsx(out_path, emails):
    """将邮件列表写入 Excel（日期、邮箱地址、回复内容）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收件箱"

    headers = ["日期", "邮件地址", "发件人", "主题", "回复内容"]
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", name="PingFang SC")
    header_fill = PatternFill("solid", fgColor="525AF7")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for item in emails:
        ws.append([
            item.get("date", ""),
            item.get("from_email", ""),
            item.get("from_name", ""),
            item.get("subject", ""),
            _truncate_excel(item.get("body", "")),
        ])

    # 列宽
    widths = [20, 32, 20, 40, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 正文列自动换行
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(str(out_path))
    wb.close()


def _truncate_excel(text, limit=32000):
    """Excel 单元格最多 32767 字符，超出截断"""
    if not text:
        return ""
    if len(text) > limit:
        return text[:limit] + "\n...(内容过长已截断)"
    return text


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def parse_excel(filepath: Path):
    """解析 xls / xlsx / csv 文件，返回 (headers, rows)"""
    ext = filepath.suffix.lower()

    if ext == ".xls":
        # 先尝试用 xlrd 解析真正的 xls 格式
        try:
            wb = xlrd.open_workbook(str(filepath))
            ws = wb.sheets()[0]
            headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
            rows = []
            for r in range(1, ws.nrows):
                row = {}
                for c in range(ws.ncols):
                    val = ws.cell_value(r, c)
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                    row[headers[c]] = str(val) if val else ""
                rows.append(row)
            return headers, rows
        except Exception:
            # 文件后缀是 .xls 但实际是 CSV/纯文本，回退到 CSV 解析
            return _parse_csv(filepath)

    elif ext == ".xlsx":
        wb = openpyxl.load_workbook(str(filepath), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []
        headers = [str(h).strip() if h else f"列{i}" for i, h in enumerate(all_rows[0])]
        rows = []
        for r in all_rows[1:]:
            row = {}
            for c, val in enumerate(r):
                if c < len(headers):
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                    row[headers[c]] = str(val) if val is not None else ""
            rows.append(row)
        wb.close()
        return headers, rows

    elif ext == ".csv":
        return _parse_csv(filepath)

    else:
        raise ValueError(f"不支持的格式: {ext}")


def _parse_csv(filepath: Path):
    """解析 CSV 或纯文本格式的文件"""
    import csv

    # 尝试检测分隔符
    with open(filepath, "r", encoding="utf-8-sig") as f:
        sample = f.read(4096)

    # 判断分隔符：制表符、逗号、或纯换行（单列）
    if "\t" in sample:
        delimiter = "\t"
    elif "," in sample:
        delimiter = ","
    else:
        delimiter = ","  # 单列数据用逗号也能正常解析

    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)
        all_lines = [row for row in reader if any(cell.strip() for cell in row)]

    if not all_lines:
        return [], []

    headers = [h.strip() if h.strip() else f"列{i}" for i, h in enumerate(all_lines[0])]
    rows = []
    for line in all_lines[1:]:
        row = {}
        for c, val in enumerate(line):
            if c < len(headers):
                row[headers[c]] = val.strip()
        if any(row.values()):
            rows.append(row)
    return headers, rows


def _render_template(template: str, data: dict) -> str:
    """将 {{列名}} 替换为对应值"""
    result = template
    for key, value in data.items():
        result = result.replace("{{" + key + "}}", str(value))
    return result


# ── 启动 ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 50)
    print("  批量邮件发送工具")
    print("  打开浏览器访问: http://localhost:5050")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5050, debug=True)
